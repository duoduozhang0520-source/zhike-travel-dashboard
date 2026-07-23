from __future__ import annotations

from collections import Counter, defaultdict
from datetime import date, datetime
import json
from pathlib import Path
from statistics import median

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parent
SOURCE = Path(
    "/Users/yiting/Library/Containers/com.bytedance.macos.feishu/Data/Downloads/"
    "26年全年分贝通-清洗数据.xlsx"
)
MONTHS = [f"{month}月" for month in range(1, 7)]
TRAVEL_TYPES = {"酒店": "酒店", "国内机票": "机票", "机票": "机票", "火车": "火车"}
ALL_TYPES = ["机票", "酒店", "火车", "用车"]


def num(value):
    try:
        return float(value or 0)
    except (TypeError, ValueError):
        return 0.0


def person_name(row):
    value = row[13] or row[9] or ""
    return str(value).strip()


def subgroup(value):
    text = str(value or "")
    if "AI课堂业务/" in text:
        return text.split("AI课堂业务/", 1)[1].split("/", 1)[0] or "其他"
    return "其他"


def date_text(value):
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, date):
        return value.isoformat()
    text = str(value or "").strip()
    return text[:10] if len(text) >= 10 else text


def load_quota(path: Path):
    result = {}
    for line in path.read_text(encoding="utf-8").splitlines():
        cols = line.split("\t")
        if len(cols) < 7 or not cols[1].startswith("AI课堂业务"):
            continue
        values = [num(value) for value in cols[2:7]]
        result[cols[0].strip()] = {
            "group": subgroup(cols[1]),
            "monthly": {
                "1月": values[0],
                "2月": values[1],
                "3月": values[2],
                "4月": values[3],
                "5月": values[4] / 3,
                "6月": values[4] / 3,
            },
        }
    return result


travel_quota = load_quota(ROOT / "travel-quota.tsv")
car_quota = load_quota(ROOT / "car-quota.tsv")

workbook = load_workbook(SOURCE, read_only=True, data_only=True)
sheet = workbook.active

raw = []
for row in sheet.iter_rows(min_row=2, values_only=True):
    department = str(row[12] or "")
    month = str(row[0] or "")
    source_type = str(row[1] or "")
    if not department.startswith("AI课堂业务") or month not in MONTHS:
        continue
    expense_type = TRAVEL_TYPES.get(source_type) or ("用车" if source_type == "用车" else None)
    if not expense_type:
        continue
    name = person_name(row)
    if not name:
        continue
    start_date = date_text(row[30]) or date_text(row[5])
    destination = str(row[28] or row[26] or "").strip()
    origin = str(row[26] or "").strip()
    status = str(row[8] or "")
    raw.append(
        {
            "month": month,
            "type": expense_type,
            "name": name,
            "group": subgroup(department),
            "amount": round(num(row[60]), 2),
            "order": str(row[2] or row[3] or ""),
            "date": start_date,
            "origin": origin,
            "destination": destination,
            "route": f"{origin}-{destination}" if origin and destination else "",
            "status": status,
            "refund": "退票" in status or "退订" in status,
            "change": "改签" in status,
            "overstandard": str(row[73] or "") == "是",
            "reason": str(row[64] or row[63] or "").strip(),
        }
    )

people_names = sorted(set(travel_quota) | set(car_quota) | {row["name"] for row in raw})
people = []
for name in people_names:
    rows = [row for row in raw if row["name"] == name]
    group = (
        travel_quota.get(name, {}).get("group")
        or car_quota.get(name, {}).get("group")
        or (rows[0]["group"] if rows else "其他")
    )
    monthly = []
    for month in MONTHS:
        month_rows = [row for row in rows if row["month"] == month]
        travel_spent = sum(row["amount"] for row in month_rows if row["type"] != "用车")
        car_spent = sum(row["amount"] for row in month_rows if row["type"] == "用车")
        tq = travel_quota.get(name, {}).get("monthly", {}).get(month, 0)
        cq = car_quota.get(name, {}).get("monthly", {}).get(month, 0)
        monthly.append(
            {
                "month": month,
                "travelSpent": round(travel_spent, 2),
                "carSpent": round(car_spent, 2),
                "spent": round(travel_spent + car_spent, 2),
                "travelQuota": round(tq, 2),
                "carQuota": round(cq, 2),
                "quota": round(tq + cq, 2),
            }
        )

    travel_spent = sum(row["amount"] for row in rows if row["type"] != "用车")
    car_spent = sum(row["amount"] for row in rows if row["type"] == "用车")
    spent = travel_spent + car_spent
    tq = sum(row["travelQuota"] for row in monthly)
    cq = sum(row["carQuota"] for row in monthly)
    quota = tq + cq
    order_count = len({row["order"] for row in rows if row["order"] and row["amount"] > 0})
    destinations = Counter(row["destination"] for row in rows if row["destination"])
    routes = Counter(row["route"] for row in rows if row["route"])
    refund_count = sum(row["refund"] for row in rows)
    change_count = sum(row["change"] for row in rows)
    over_count = sum(row["overstandard"] for row in rows)
    people.append(
        {
            "name": name,
            "group": group,
            "travelSpent": round(travel_spent, 2),
            "carSpent": round(car_spent, 2),
            "spent": round(spent, 2),
            "travelQuota": round(tq, 2),
            "carQuota": round(cq, 2),
            "quota": round(quota, 2),
            "rate": round(spent / quota, 6) if quota else None,
            "orderCount": order_count,
            "avgOrder": round(spent / order_count, 2) if order_count else 0,
            "refundCount": refund_count,
            "changeCount": change_count,
            "overstandardCount": over_count,
            "topCities": [item[0] for item in destinations.most_common(5)],
            "topRoutes": [item[0] for item in routes.most_common(4)],
            "monthly": monthly,
            "labels": [],
            "score": 0,
        }
    )

group_medians = {}
for group in sorted({person["group"] for person in people}):
    values = [person["spent"] for person in people if person["group"] == group and person["spent"] > 0]
    group_medians[group] = median(values) if values else 0

for person in people:
    labels = []
    score = 0
    if person["quota"] == 0 and person["spent"] > 0:
        labels.append("无额度但有消费")
        score += 5
    elif person["rate"] is not None and person["rate"] >= 1:
        labels.append("累计额度超额")
        score += 5
    elif person["rate"] is not None and person["rate"] >= 0.8:
        labels.append("累计额度预警")
        score += 3

    latest = person["monthly"][-1]["spent"]
    history = [item["spent"] for item in person["monthly"][:-1] if item["spent"] > 0]
    history_avg = sum(history) / len(history) if history else 0
    if latest > 1500 and history_avg and latest >= history_avg * 1.8:
        labels.append("6月显著上升")
        score += 2

    peer_median = group_medians.get(person["group"], 0)
    if person["spent"] > 3000 and peer_median and person["spent"] >= peer_median * 1.8:
        labels.append("高于同组中位数")
        score += 2
    if person["refundCount"] >= 3:
        labels.append("退票/退订较多")
        score += 2
    if person["changeCount"] >= 3:
        labels.append("改签较多")
        score += 1
    if person["overstandardCount"] >= 3:
        labels.append("超规记录较多")
        score += 2
    if len(person["topCities"]) >= 5:
        labels.append("出行城市分散")
        score += 1
    person["labels"] = labels
    person["score"] = score
    person["peerMedian"] = round(peer_median, 2)

groups = []
for group in sorted({person["group"] for person in people}):
    members = [person for person in people if person["group"] == group]
    spenders = [person for person in members if person["spent"] != 0]
    spent = sum(person["spent"] for person in members)
    tq = sum(person["travelQuota"] for person in members)
    cq = sum(person["carQuota"] for person in members)
    quota = tq + cq
    warning_people = sum(person["score"] >= 3 for person in members)
    groups.append(
        {
            "group": group,
            "headcount": sum(person["quota"] > 0 for person in members),
            "spenderCount": len(spenders),
            "spent": round(spent, 2),
            "travelSpent": round(sum(person["travelSpent"] for person in members), 2),
            "carSpent": round(sum(person["carSpent"] for person in members), 2),
            "quota": round(quota, 2),
            "travelQuota": round(tq, 2),
            "carQuota": round(cq, 2),
            "rate": round(spent / quota, 6) if quota else None,
            "perHead": round(spent / max(sum(person["quota"] > 0 for person in members), 1), 2),
            "perSpender": round(spent / len(spenders), 2) if spenders else 0,
            "warningPeople": warning_people,
            "warningRate": round(warning_people / len(members), 6) if members else 0,
            "topPeople": [person["name"] for person in sorted(members, key=lambda item: item["spent"], reverse=True)[:3]],
        }
    )

monthly = []
for month in MONTHS:
    rows = [row for row in raw if row["month"] == month]
    month_people = [person for person in people if any(item["month"] == month and item["spent"] != 0 for item in person["monthly"])]
    item = {
        "month": month,
        "spent": round(sum(row["amount"] for row in rows), 2),
        "spenders": len(month_people),
        "orders": len({row["order"] for row in rows if row["order"] and row["amount"] > 0}),
    }
    for expense_type in ALL_TYPES:
        item[expense_type] = round(sum(row["amount"] for row in rows if row["type"] == expense_type), 2)
    item["quota"] = round(
        sum(person_month["quota"] for person in people for person_month in person["monthly"] if person_month["month"] == month),
        2,
    )
    item["perSpender"] = round(item["spent"] / item["spenders"], 2) if item["spenders"] else 0
    item["avgOrder"] = round(item["spent"] / item["orders"], 2) if item["orders"] else 0
    monthly.append(item)

event_groups = defaultdict(list)
for row in raw:
    if row["date"] and row["destination"] and row["amount"] > 0:
        event_groups[(row["date"], row["destination"])].append(row)
events = []
for (event_date, destination), rows in event_groups.items():
    names = sorted({row["name"] for row in rows})
    if len(names) < 3:
        continue
    events.append(
        {
            "date": event_date,
            "destination": destination,
            "people": names,
            "count": len(names),
            "amount": round(sum(row["amount"] for row in rows), 2),
            "groups": sorted({row["group"] for row in rows}),
        }
    )
events.sort(key=lambda item: (item["count"], item["amount"]), reverse=True)

summary = {
    "travelSpent": round(sum(person["travelSpent"] for person in people), 2),
    "carSpent": round(sum(person["carSpent"] for person in people), 2),
    "spent": round(sum(person["spent"] for person in people), 2),
    "travelQuota": round(sum(person["travelQuota"] for person in people), 2),
    "carQuota": round(sum(person["carQuota"] for person in people), 2),
    "quota": round(sum(person["quota"] for person in people), 2),
    "headcount": sum(person["quota"] > 0 for person in people),
    "spenderCount": sum(person["spent"] != 0 for person in people),
    "warningPeople": sum(person["score"] >= 3 for person in people),
    "overPeople": sum("累计额度超额" in person["labels"] for person in people),
    "noQuotaPeople": sum("无额度但有消费" in person["labels"] for person in people),
}
summary["rate"] = round(summary["spent"] / summary["quota"], 6) if summary["quota"] else 0
summary["perHead"] = round(summary["spent"] / summary["headcount"], 2) if summary["headcount"] else 0
summary["perSpender"] = round(summary["spent"] / summary["spenderCount"], 2) if summary["spenderCount"] else 0

output = {
    "title": "AI课堂业务部-分贝通看板",
    "generatedAt": datetime.now().strftime("%Y-%m-%d %H:%M"),
    "period": "2026年1–6月",
    "sourceRows": len(raw),
    "summary": summary,
    "monthly": monthly,
    "groups": sorted(groups, key=lambda item: item["spent"], reverse=True),
    "people": sorted(people, key=lambda item: (-item["score"], -item["spent"])),
    "events": events[:20],
}

(ROOT / "ai-data.json").write_text(
    json.dumps(output, ensure_ascii=False, separators=(",", ":")),
    encoding="utf-8",
)
print(json.dumps(summary, ensure_ascii=False, indent=2))
print(f"AI records: {len(raw)}, people: {len(people)}, event clusters: {len(events)}")
